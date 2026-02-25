#!/usr/bin/env python3
"""
Download revisions from Google Sheets and extract changes in
Situação, Etapa, Planejamento Admissão columns.
Uses composite key (Cargo+Destino+Gestor) to identify each vaga.
Handles two column formats:
  - Old (rev 1-9, 11): Cargo, Destino, Gestor, E-mail, Situação, Etapa, Previsão Admissão, ...
  - New (rev 10, 12+): id, Recrutador, Cargo, Destino, Gestor, E-mail, Situação, Etapa, Planejamento Admissão, ...
"""

import subprocess
import os
import json
import time
import sys
import re
from datetime import datetime

try:
    from openpyxl import load_workbook
except ImportError:
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl'])
    from openpyxl import load_workbook

import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

COOKIES = open('/tmp/google_cookies.txt').read().strip()
SHEET_ID = '1aW71mfAHBSbzIeS9-be9qcpfzTWjzEPf'
DOWNLOAD_DIR = '/tmp/ug_revisions'
os.makedirs(DOWNLOAD_DIR, exist_ok=True)


def download_revision(rev_num):
    """Download a revision as XLSX, return filepath or None."""
    filepath = os.path.join(DOWNLOAD_DIR, f'rev_{rev_num}.xlsx')
    if os.path.exists(filepath) and os.path.getsize(filepath) > 1000:
        return filepath

    url = f'https://docs.google.com/spreadsheets/export?id={SHEET_ID}&revision={rev_num}&exportFormat=xlsx'
    result = subprocess.run(
        ['curl', '-sL', '-w', '%{http_code}', '-o', filepath, '-b', COOKIES, url],
        capture_output=True, text=True, timeout=60
    )
    http_code = result.stdout.strip()

    if http_code == '200' and os.path.exists(filepath) and os.path.getsize(filepath) > 1000:
        return filepath
    else:
        if os.path.exists(filepath):
            os.remove(filepath)
        return None


def parse_sheet(filepath):
    """Parse XLSX and extract data for each row.
    Returns dict keyed by composite key (Cargo|Destino|Gestor) with values dict.
    """
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        return None

    # Find Acompanhamento sheet
    sheet = None
    for name in wb.sheetnames:
        if 'acompanhamento' in name.lower():
            sheet = wb[name]
            break
    if sheet is None:
        sheet = wb[wb.sheetnames[0]]

    # Read header row and auto-detect columns
    headers = []
    for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = [str(h).strip().lower() if h else '' for h in row]

    if not headers:
        wb.close()
        return None

    # Build column map by header name
    col_map = {}
    for i, h in enumerate(headers):
        if h == 'id':
            col_map['id'] = i
        elif h.startswith('recrutador'):
            col_map['Recrutador'] = i
        elif h == 'cargo':
            col_map['Cargo'] = i
        elif h == 'destino':
            col_map['Destino'] = i
        elif h == 'gestor':
            col_map['Gestor'] = i
        elif h.startswith('situa'):
            col_map['Situacao'] = i
        elif h == 'etapa':
            col_map['Etapa'] = i
        elif h.startswith('planejamento') or h.startswith('previs'):
            col_map['Planejamento Admissao'] = i

    # Verify we have the essential columns
    required = ['Cargo', 'Destino', 'Gestor', 'Situacao', 'Etapa', 'Planejamento Admissao']
    for req in required:
        if req not in col_map:
            wb.close()
            return None

    # Extract data
    rows = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        try:
            def get_val(col_name):
                idx = col_map.get(col_name)
                if idx is not None and idx < len(row):
                    val = row[idx]
                    if val is not None:
                        s = str(val).strip()
                        return '' if s.lower() == 'none' else s
                return ''

            cargo = get_val('Cargo')
            destino = get_val('Destino')
            gestor = get_val('Gestor')

            # Skip empty rows
            if not cargo:
                continue

            # Composite key: Cargo|Destino|Gestor
            composite_key = f"{cargo}|{destino}|{gestor}"

            id_val = get_val('id')
            recrutador = get_val('Recrutador')
            situacao = get_val('Situacao')
            etapa = get_val('Etapa')
            planejamento = get_val('Planejamento Admissao')

            rows[composite_key] = {
                'id': id_val,
                'Recrutador': recrutador,
                'Cargo': cargo,
                'Destino': destino,
                'Gestor': gestor,
                'Situacao': situacao,
                'Etapa': etapa,
                'Planejamento Admissao': planejamento
            }
        except (IndexError, TypeError):
            continue

    wb.close()
    return rows


def parse_date(date_str):
    """Parse Portuguese date string to datetime."""
    months = {
        'janeiro': 1, 'fevereiro': 2, 'março': 3, 'abril': 4,
        'maio': 5, 'junho': 6, 'julho': 7, 'agosto': 8,
        'setembro': 9, 'outubro': 10, 'novembro': 11, 'dezembro': 12
    }
    m = re.match(r'(\d+) de (\w+)(?:\s+de\s+(\d{4}))?,?\s*(\d+):(\d+)', date_str)
    if m:
        day = int(m.group(1))
        month = months.get(m.group(2).lower(), 1)
        year = int(m.group(3)) if m.group(3) else 2026
        hour = int(m.group(4))
        minute = int(m.group(5))
        # If month > February in 2026, it's from 2025
        if year == 2026 and month > 2:
            year = 2025
        return datetime(year, month, day, hour, minute)
    return None


def main():
    # Load revision map
    with open('/tmp/ug_revision_map.json') as f:
        revisions = json.load(f)

    valid_revs = [r for r in revisions if r.get('rev') is not None]
    valid_revs.sort(key=lambda x: x['rev'])
    print(f"Total revisions: {len(revisions)}, with rev ID: {len(valid_revs)}")

    all_changes = []
    prev_data = None
    prev_rev = None
    prev_keys = set()
    skipped = 0

    for i, rev_info in enumerate(valid_revs):
        rev = rev_info['rev']
        date_str = rev_info['date']
        authors = rev_info.get('authors', [])
        author_str = ', '.join(authors) if authors else 'Desconhecido'

        print(f"[{i+1}/{len(valid_revs)}] Rev {rev} - {date_str} ({author_str})...", end=' ', flush=True)

        filepath = download_revision(rev)
        if not filepath:
            for retry in range(3):
                wait = 5 * (retry + 1)
                print(f"RETRY({retry+1})...", end=' ', flush=True)
                time.sleep(wait)
                filepath = download_revision(rev)
                if filepath:
                    break
            if not filepath:
                print("SKIPPED")
                skipped += 1
                continue

        data = parse_sheet(filepath)
        if not data:
            print("PARSE ERROR")
            continue

        current_keys = set(data.keys())
        parsed_date = parse_date(date_str)
        date_iso = parsed_date.isoformat() if parsed_date else date_str

        changes_count = 0

        # Detect NEW vagas (inclusão de vaga)
        if prev_data is not None:
            new_keys = current_keys - prev_keys
            for key in new_keys:
                vaga = data[key]
                all_changes.append({
                    'composite_key': key,
                    'id': vaga['id'],
                    'Recrutador': vaga['Recrutador'],
                    'Cargo': vaga['Cargo'],
                    'Destino': vaga['Destino'],
                    'Gestor': vaga['Gestor'],
                    'coluna_alterada': 'Inclusão de Vaga',
                    'data_alteracao': date_iso,
                    'Situacao': vaga['Situacao'],
                    'Etapa': vaga['Etapa'],
                    'Planejamento Admissao': vaga['Planejamento Admissao'],
                    'alterado_por': author_str,
                    'rev': rev,
                    'prev_rev': prev_rev
                })
                changes_count += 1

        # Detect changes in tracked columns
        if prev_data is not None:
            for key, current in data.items():
                prev = prev_data.get(key)
                if not prev:
                    continue  # Already handled as new vaga above

                for col_name in ['Situacao', 'Etapa', 'Planejamento Admissao']:
                    current_val = current.get(col_name, '')
                    prev_val = prev.get(col_name, '')
                    if current_val != prev_val:
                        all_changes.append({
                            'composite_key': key,
                            'id': current['id'],
                            'Recrutador': current['Recrutador'],
                            'Cargo': current['Cargo'],
                            'Destino': current['Destino'],
                            'Gestor': current['Gestor'],
                            'coluna_alterada': col_name,
                            'data_alteracao': date_iso,
                            'Situacao': current['Situacao'],
                            'Etapa': current['Etapa'],
                            'Planejamento Admissao': current['Planejamento Admissao'],
                            'alterado_por': author_str,
                            'rev': rev,
                            'prev_rev': prev_rev
                        })
                        changes_count += 1

        if prev_data is None:
            # First revision: all vagas are "inclusão"
            for key, vaga in data.items():
                all_changes.append({
                    'composite_key': key,
                    'id': vaga['id'],
                    'Recrutador': vaga['Recrutador'],
                    'Cargo': vaga['Cargo'],
                    'Destino': vaga['Destino'],
                    'Gestor': vaga['Gestor'],
                    'coluna_alterada': 'Inclusão de Vaga',
                    'data_alteracao': date_iso,
                    'Situacao': vaga['Situacao'],
                    'Etapa': vaga['Etapa'],
                    'Planejamento Admissao': vaga['Planejamento Admissao'],
                    'alterado_por': author_str,
                    'rev': rev,
                    'prev_rev': None
                })
                changes_count += 1
            print(f"{len(data)} rows | {changes_count} initial entries")
        else:
            print(f"{len(data)} rows | {changes_count} changes")

        prev_data = data
        prev_rev = rev
        prev_keys = current_keys

        # Save progress
        if i % 20 == 19:
            with open('/tmp/ug_changes_progress.json', 'w') as f:
                json.dump(all_changes, f, indent=2, ensure_ascii=False)
            print(f"  ... saved ({len(all_changes)} total)")

        time.sleep(2)

    # Save final results
    with open('/tmp/ug_changes_full.json', 'w') as f:
        json.dump(all_changes, f, indent=2, ensure_ascii=False)

    print(f"\n=== DONE ===")
    print(f"Total entries: {len(all_changes)}")
    print(f"Skipped revisions: {skipped}")

    col_counts = {}
    for c in all_changes:
        col = c['coluna_alterada']
        col_counts[col] = col_counts.get(col, 0) + 1
    print(f"\nBy type:")
    for col, count in sorted(col_counts.items(), key=lambda x: -x[1]):
        print(f"  {col}: {count}")

    vaga_counts = {}
    for c in all_changes:
        vaga = c['Cargo'][:40]
        vaga_counts[vaga] = vaga_counts.get(vaga, 0) + 1
    print(f"\nTop vagas:")
    for vaga, count in sorted(vaga_counts.items(), key=lambda x: -x[1])[:15]:
        print(f"  {vaga}: {count}")


if __name__ == '__main__':
    main()
