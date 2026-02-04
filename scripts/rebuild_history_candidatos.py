#!/usr/bin/env python3
"""
Complete History Rebuild for Candidatos from Google Sheets Version History.

Creates PARALLEL table log_history_rebuildCandidatos (does NOT modify USO_historicoCandidatos).
This table has the same structure as USO_historicoCandidatos and is populated as if
the AppScript + API had been running since January 2025.

LOGIC (based on GATILHO_mudanca_status_candidato.gs):
- When status_candidato OR status_micro_candidato changes → Create new row in historico

COLUMN MAPPING (Excel position → PostgreSQL field):
  0  (ID do candidato)      → id_candidato
  1  (Nome)                 → nome_candidato
  2  (Linkedin)             → linkedin_candidato
  3  (Telefone/ E-mail)     → contatos_candidato / telefone_candidato
  4  (Localidade candidato) → localidade_candidato
  5  (Canal de atratividade)→ candidato_origem
  8  (Status Candidato)     → status_candidato
  9  (Status Candidato micro)→ status_micro_candidato
"""

import json
import os
import re
from datetime import datetime
from openpyxl import load_workbook
import psycopg2

# Database connection
DB_CONFIG = {
    'host': 'localhost',
    'port': 5432,
    'database': 'HUB',
    'user': 'rpo_user',
    'password': 'rpo_super_secret001'
}

# Directory with downloaded revision files
REVISION_DIR = '/tmp/sheets_revisions'

# Map short names to emails
EMAIL_MAP = {
    'Ariane': 'ariane.moura@hubtalent.com.br',
    'Carolina': 'carolina.silva@hubinfinity.com.br',
    'Cazou': 'cazou.vilela@hubtalent.com.br',
    'Emily': 'emily.alencar@hubinfinity.com.br',
    'Gabrielle': 'gabrielle@hubinfinity.com.br',
    'Gessica': 'gessica.amorim@hubtalent.com.br',
    'Jessica': 'jessica.morais@hubinfinity.com.br',
    'Liliane': 'liliane.ferreira@hubinfinity.com.br',
    'Taina': 'taina.lustosa@hubinfinity.com.br'
}

# Column indices (0-based) in the Excel file
COLUMNS = {
    'id_candidato': 0,
    'nome_candidato': 1,
    'linkedin_candidato': 2,
    'contatos_candidato': 3,
    'localidade_candidato': 4,
    'candidato_origem': 5,
    'status_candidato': 8,
    'status_micro_candidato': 9,
}


def parse_date(date_str, year=2025):
    """Parse Portuguese date string to datetime."""
    months = {
        'janeiro': 1, 'fevereiro': 2, 'março': 3, 'abril': 4,
        'maio': 5, 'junho': 6, 'julho': 7, 'agosto': 8,
        'setembro': 9, 'outubro': 10, 'novembro': 11, 'dezembro': 12
    }
    m = re.match(r'(\d+) de (\w+),? (\d+):(\d+)', date_str)
    if m:
        day = int(m.group(1))
        month = months.get(m.group(2).lower(), 1)
        hour = int(m.group(3))
        minute = int(m.group(4))
        return datetime(year, month, day, hour, minute)
    return None


def get_cell_value(row, col_idx):
    """Safely get cell value from row."""
    if col_idx < len(row):
        val = row[col_idx]
        if val is None:
            return None
        val_str = str(val).strip()
        if val_str.lower() == 'none' or val_str == '':
            return None
        if val_str.endswith('.0'):
            val_str = val_str[:-2]
        return val_str
    return None


def parse_candidatos_sheet(filepath):
    """
    Parse XLSX and extract ALL relevant fields for each candidato.
    Returns dict: {id_candidato: {field1: value1, field2: value2, ...}}
    """
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        return None

    # Find candidatos sheet
    sheet = None
    for name in wb.sheetnames:
        if 'candidato' in name.lower():
            sheet = wb[name]
            break
    if sheet is None:
        if len(wb.sheetnames) >= 3:
            sheet = wb[wb.sheetnames[2]]
        else:
            wb.close()
            return None

    # Extract data
    rows = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        id_cand = get_cell_value(row, COLUMNS['id_candidato'])
        if not id_cand:
            continue

        row_data = {}
        for field, col_idx in COLUMNS.items():
            row_data[field] = get_cell_value(row, col_idx)

        rows[id_cand] = row_data

    wb.close()
    return rows


def main():
    print("=== Complete History Rebuild for Candidatos ===\n")

    # Load revision map
    with open('/tmp/revision_map_full.json') as f:
        revisions = json.load(f)

    # Filter to human-edited revisions
    human_revs = [r for r in revisions if r.get('hasHuman') and r.get('rev')]
    human_revs_sorted = sorted(human_revs, key=lambda x: x['rev'])

    print(f"Total revisions: {len(revisions)}")
    print(f"Human-edited with rev ID: {len(human_revs)}\n")

    # Connect to database
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # Create table
    print("Creating log_history_rebuildCandidatos table...")
    cur.execute("""
        DROP TABLE IF EXISTS "RPO_cielo"."log_history_rebuildCandidatos";

        CREATE TABLE "RPO_cielo"."log_history_rebuildCandidatos" (
            id SERIAL PRIMARY KEY,
            id_candidato VARCHAR(100) NOT NULL,
            status_candidato VARCHAR(200),
            status_micro_candidato VARCHAR(200),
            alterado_por VARCHAR(255),
            created_at TIMESTAMP WITHOUT TIME ZONE NOT NULL,
            updated_at TIMESTAMP WITHOUT TIME ZONE NOT NULL,
            nome_candidato TEXT,
            contatos_candidato TEXT,
            localidade_candidato TEXT,
            telefone_candidato TEXT,
            vaga_salario TEXT,
            candidato_origem TEXT,
            linkedin_candidato TEXT
        );

        CREATE INDEX idx_log_history_rebuild_cand_id ON "RPO_cielo"."log_history_rebuildCandidatos"(id_candidato);
        CREATE INDEX idx_log_history_rebuild_cand_created ON "RPO_cielo"."log_history_rebuildCandidatos"(created_at DESC);
    """)
    conn.commit()
    print("Table created.\n")

    # Process revisions
    prev_data = None
    history_records = []
    processed = 0
    skipped = 0

    for i, rev_info in enumerate(human_revs_sorted):
        rev = rev_info['rev']
        date_str = rev_info['date']
        human = rev_info.get('human', 'Unknown')
        change_datetime = parse_date(date_str)

        filepath = os.path.join(REVISION_DIR, f'rev_{rev}.xlsx')
        if not os.path.exists(filepath):
            skipped += 1
            continue

        data = parse_candidatos_sheet(filepath)
        if not data:
            skipped += 1
            continue

        # Compare with previous revision
        if prev_data is not None:
            for id_cand, current_row in data.items():
                current_status = current_row.get('status_candidato', '')
                current_micro = current_row.get('status_micro_candidato', '')

                prev_row = prev_data.get(id_cand, {})
                prev_status = prev_row.get('status_candidato', '')
                prev_micro = prev_row.get('status_micro_candidato', '')

                # Create history record if status OR micro status changed
                status_changed = current_status and current_status != prev_status
                micro_changed = current_micro and current_micro != prev_micro

                if status_changed or micro_changed:
                    email = EMAIL_MAP.get(human, f'{human.lower()}@hubinfinity.com.br')

                    record = {
                        'id_candidato': id_cand,
                        'status_candidato': current_status,
                        'status_micro_candidato': current_micro,
                        'alterado_por': email,
                        'created_at': change_datetime,
                        'updated_at': change_datetime,
                        'nome_candidato': current_row.get('nome_candidato'),
                        'contatos_candidato': current_row.get('contatos_candidato'),
                        'localidade_candidato': current_row.get('localidade_candidato'),
                        'candidato_origem': current_row.get('candidato_origem'),
                        'linkedin_candidato': current_row.get('linkedin_candidato'),
                    }
                    history_records.append(record)

        prev_data = data
        processed += 1

        if processed % 20 == 0:
            print(f"  Processed {processed} revisions, {len(history_records)} history records...")

    print(f"\nProcessed {processed} revisions, skipped {skipped}")
    print(f"Total history records to insert: {len(history_records)}\n")

    # Insert records
    print("Inserting records into log_history_rebuildCandidatos...")

    insert_sql = """
        INSERT INTO "RPO_cielo"."log_history_rebuildCandidatos"
        (id_candidato, status_candidato, status_micro_candidato, alterado_por,
         created_at, updated_at, nome_candidato, contatos_candidato,
         localidade_candidato, candidato_origem, linkedin_candidato)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """

    for record in history_records:
        cur.execute(insert_sql, (
            record['id_candidato'],
            record['status_candidato'],
            record['status_micro_candidato'],
            record['alterado_por'],
            record['created_at'],
            record['updated_at'],
            record['nome_candidato'],
            record['contatos_candidato'],
            record['localidade_candidato'],
            record['candidato_origem'],
            record['linkedin_candidato'],
        ))

    conn.commit()
    print(f"Inserted {len(history_records)} records.\n")

    # Summary
    print("=== Summary ===")

    cur.execute('SELECT COUNT(*) FROM "RPO_cielo"."log_history_rebuildCandidatos"')
    total = cur.fetchone()[0]
    print(f"Total rows: {total}")

    cur.execute('SELECT COUNT(DISTINCT id_candidato) FROM "RPO_cielo"."log_history_rebuildCandidatos"')
    unique_cands = cur.fetchone()[0]
    print(f"Unique candidatos: {unique_cands}")

    # Status distribution
    print("\nStatus candidato distribution:")
    cur.execute("""
        SELECT status_candidato, COUNT(*) as cnt
        FROM "RPO_cielo"."log_history_rebuildCandidatos"
        WHERE status_candidato IS NOT NULL
        GROUP BY status_candidato
        ORDER BY cnt DESC
        LIMIT 15
    """)
    for status, cnt in cur.fetchall():
        print(f"  {status}: {cnt}")

    # Micro status distribution
    print("\nStatus micro candidato distribution:")
    cur.execute("""
        SELECT status_micro_candidato, COUNT(*) as cnt
        FROM "RPO_cielo"."log_history_rebuildCandidatos"
        WHERE status_micro_candidato IS NOT NULL
        GROUP BY status_micro_candidato
        ORDER BY cnt DESC
        LIMIT 10
    """)
    for status, cnt in cur.fetchall():
        print(f"  {status}: {cnt}")

    # Sample records
    print("\nSample records:")
    cur.execute("""
        SELECT id_candidato, nome_candidato, status_candidato,
               status_micro_candidato, created_at
        FROM "RPO_cielo"."log_history_rebuildCandidatos"
        ORDER BY created_at
        LIMIT 5
    """)
    for row in cur.fetchall():
        print(f"  {row[0]}: {row[1]} | {row[2]} | {row[3]} | {row[4]}")

    # Compare with USO_historicoCandidatos
    print("\n=== Comparison with USO_historicoCandidatos ===")
    cur.execute('SELECT COUNT(*) FROM "RPO_cielo"."USO_historicoCandidatos"')
    uso_count = cur.fetchone()[0]
    print(f"USO_historicoCandidatos rows: {uso_count}")
    print(f"log_history_rebuildCandidatos rows: {total}")

    cur.close()
    conn.close()

    print("\n=== Done! ===")


if __name__ == '__main__':
    main()
