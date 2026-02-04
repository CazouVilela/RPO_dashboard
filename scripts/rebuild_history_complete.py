#!/usr/bin/env python3
"""
Complete History Rebuild from Google Sheets Version History.

Creates PARALLEL table log_history_rebuild (does NOT modify USO_historicoVagas).
This table has the same structure as USO_historicoVagas and is populated as if
the AppScript + API had been running since January 2025.

LOGIC (based on GATILHO_mudanca_status.gs and EFEITO_inclui_API_HISTORICO.gs):
1. When STATUS changes → Create new row in historico with ALL current field values:
   - requisicao, status, alterado_por, created_at
   - DADOS_CANDIDATOS: cdd_totais_sl, cdd_mulher_sl, cdd_pcd_sl, cdd_diversidade_racial_sl
   - DADOS_SELECIONADO: selecionado_nome, selecionado_fonte, selecionado_genero,
     selecionado_pcd, selecionado_diversidade_racial, selecionado_tipo,
     selecionado_empresa_anterior, selecionado_empregado
   - DADOS_PROPOSTA: operacao_status_ultima_proposta, operacao_selecionado_ultima_proposta,
     operacao_motivo_declinio_ultima_proposta

COLUMN MAPPING (Excel position → PostgreSQL field):
  5  (RP)                               → requisicao
  19 (Status)                           → status
  24 (Total de candidatos na shortlist) → cdd_totais_sl
  25 (Mulheres na shortlist)            → cdd_mulher_sl
  26 (PCDs na shortlist)                → cdd_pcd_sl
  27 (Diversidade racial na shortlist)  → cdd_diversidade_racial_sl
  28 (Nome do selecionado)              → selecionado_nome
  29 (Origem do selecionado)            → selecionado_fonte
  30 (Gênero do selecionado)            → selecionado_genero
  31 (Selecionado é PCD?)               → selecionado_pcd
  32 (Raça do selecionado)              → selecionado_diversidade_racial
  33 (Segmento do selecionado)          → selecionado_tipo
  34 (Empresa anterior do selecionado)  → selecionado_empresa_anterior
  35 (Selecionado estava empregado?)    → selecionado_empregado
  37 (Status da ultima proposta)        → operacao_status_ultima_proposta
  38 (Selecionado da ultima proposta)   → operacao_selecionado_ultima_proposta
  39 (Motivo do declínio)               → operacao_motivo_declinio_ultima_proposta
"""

import json
import os
import re
from datetime import datetime
from openpyxl import load_workbook
import psycopg2

# Database connection
DB_CONFIG = {
    'host': 'localhost',
    'port': 5432,
    'database': 'HUB',
    'user': 'rpo_user',
    'password': 'rpo_super_secret001'
}

# Directory with downloaded revision files
REVISION_DIR = '/tmp/sheets_revisions'

# Map short names to emails
EMAIL_MAP = {
    'Ariane': 'ariane.moura@hubtalent.com.br',
    'Carolina': 'carolina.silva@hubinfinity.com.br',
    'Cazou': 'cazou.vilela@hubtalent.com.br',
    'Emily': 'emily.alencar@hubinfinity.com.br',
    'Gabrielle': 'gabrielle@hubinfinity.com.br',
    'Gessica': 'gessica.amorim@hubtalent.com.br',
    'Jessica': 'jessica.morais@hubinfinity.com.br',
    'Liliane': 'liliane.ferreira@hubinfinity.com.br',
    'Taina': 'taina.lustosa@hubinfinity.com.br'
}

# Column indices (0-based) in the Excel file
COLUMNS = {
    'rp': 5,
    'status': 19,
    'cdd_totais_sl': 24,
    'cdd_mulher_sl': 25,
    'cdd_pcd_sl': 26,
    'cdd_diversidade_racial_sl': 27,
    'selecionado_nome': 28,
    'selecionado_fonte': 29,
    'selecionado_genero': 30,
    'selecionado_pcd': 31,
    'selecionado_diversidade_racial': 32,
    'selecionado_tipo': 33,
    'selecionado_empresa_anterior': 34,
    'selecionado_empregado': 35,
    'operacao_status_ultima_proposta': 37,
    'operacao_selecionado_ultima_proposta': 38,
    'operacao_motivo_declinio_ultima_proposta': 39
}


def parse_date(date_str, year=2025):
    """Parse Portuguese date string to datetime."""
    months = {
        'janeiro': 1, 'fevereiro': 2, 'março': 3, 'abril': 4,
        'maio': 5, 'junho': 6, 'julho': 7, 'agosto': 8,
        'setembro': 9, 'outubro': 10, 'novembro': 11, 'dezembro': 12
    }
    m = re.match(r'(\d+) de (\w+),? (\d+):(\d+)', date_str)
    if m:
        day = int(m.group(1))
        month = months.get(m.group(2).lower(), 1)
        hour = int(m.group(3))
        minute = int(m.group(4))
        return datetime(year, month, day, hour, minute)
    return None


def get_cell_value(row, col_idx):
    """Safely get cell value from row."""
    if col_idx < len(row):
        val = row[col_idx]
        if val is None:
            return None
        val_str = str(val).strip()
        if val_str.lower() == 'none' or val_str == '':
            return None
        # Remove .0 suffix from numbers
        if val_str.endswith('.0'):
            val_str = val_str[:-2]
        return val_str
    return None


def parse_vagas_sheet_complete(filepath):
    """
    Parse XLSX and extract ALL relevant fields for each row in vagas sheet.
    Returns dict: {rp: {field1: value1, field2: value2, ...}}
    """
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        print(f"    Error loading {filepath}: {e}")
        return None

    # Find vagas sheet
    sheet = None
    for name in wb.sheetnames:
        if 'vagas' == name.lower():
            sheet = wb[name]
            break
    if sheet is None:
        if len(wb.sheetnames) >= 2:
            sheet = wb[wb.sheetnames[1]]
        else:
            wb.close()
            return None

    # Extract data
    rows = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        rp = get_cell_value(row, COLUMNS['rp'])
        if not rp:
            continue

        row_data = {}
        for field, col_idx in COLUMNS.items():
            row_data[field] = get_cell_value(row, col_idx)

        rows[rp] = row_data

    wb.close()
    return rows


def main():
    print("=== Complete History Rebuild from Version History ===\n")

    # Load revision map
    with open('/tmp/revision_map_full.json') as f:
        revisions = json.load(f)

    # Filter to human-edited revisions
    human_revs = [r for r in revisions if r.get('hasHuman') and r.get('rev')]
    human_revs_sorted = sorted(human_revs, key=lambda x: x['rev'])

    print(f"Total revisions: {len(revisions)}")
    print(f"Human-edited with rev ID: {len(human_revs)}\n")

    # Connect to database
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # Create table
    print("Creating log_history_rebuild table...")
    cur.execute("""
        DROP TABLE IF EXISTS "RPO_cielo".log_history_rebuild;

        CREATE TABLE "RPO_cielo".log_history_rebuild (
            id SERIAL PRIMARY KEY,
            requisicao VARCHAR(100) NOT NULL,
            status VARCHAR(100) NOT NULL,
            alterado_por VARCHAR(255),
            created_at TIMESTAMP WITHOUT TIME ZONE NOT NULL,
            updated_at TIMESTAMP WITHOUT TIME ZONE NOT NULL,
            version INTEGER DEFAULT 1,
            cdd_totais_sl TEXT,
            cdd_mulher_sl TEXT,
            cdd_pcd_sl TEXT,
            cdd_diversidade_racial_sl TEXT,
            selecionado_nome TEXT,
            selecionado_fonte TEXT,
            selecionado_genero TEXT,
            selecionado_pcd TEXT,
            selecionado_diversidade_racial TEXT,
            operacao_status_ultima_proposta TEXT,
            operacao_selecionado_ultima_proposta TEXT,
            operacao_motivo_declinio_ultima_proposta TEXT,
            selecionado_tipo TEXT,
            selecionado_empregado TEXT,
            selecionado_empresa_anterior TEXT
        );

        CREATE INDEX idx_log_history_rebuild_req ON "RPO_cielo".log_history_rebuild(requisicao);
        CREATE INDEX idx_log_history_rebuild_created ON "RPO_cielo".log_history_rebuild(created_at DESC);
        CREATE INDEX idx_log_history_rebuild_status ON "RPO_cielo".log_history_rebuild(status);
    """)
    conn.commit()
    print("Table created.\n")

    # Process revisions
    prev_data = None
    history_records = []
    processed = 0
    skipped = 0

    for i, rev_info in enumerate(human_revs_sorted):
        rev = rev_info['rev']
        date_str = rev_info['date']
        human = rev_info.get('human', 'Unknown')
        change_datetime = parse_date(date_str)

        filepath = os.path.join(REVISION_DIR, f'rev_{rev}.xlsx')
        if not os.path.exists(filepath):
            skipped += 1
            continue

        data = parse_vagas_sheet_complete(filepath)
        if not data:
            skipped += 1
            continue

        # Compare with previous revision
        if prev_data is not None:
            for rp, current_row in data.items():
                current_status = current_row.get('status', '')
                prev_row = prev_data.get(rp, {})
                prev_status = prev_row.get('status', '')

                # Only create history record if STATUS changed
                if current_status and current_status != prev_status:
                    email = EMAIL_MAP.get(human, f'{human.lower()}@hubinfinity.com.br')

                    record = {
                        'requisicao': rp,
                        'status': current_status,
                        'alterado_por': email,
                        'created_at': change_datetime,
                        'updated_at': change_datetime,
                        'cdd_totais_sl': current_row.get('cdd_totais_sl'),
                        'cdd_mulher_sl': current_row.get('cdd_mulher_sl'),
                        'cdd_pcd_sl': current_row.get('cdd_pcd_sl'),
                        'cdd_diversidade_racial_sl': current_row.get('cdd_diversidade_racial_sl'),
                        'selecionado_nome': current_row.get('selecionado_nome'),
                        'selecionado_fonte': current_row.get('selecionado_fonte'),
                        'selecionado_genero': current_row.get('selecionado_genero'),
                        'selecionado_pcd': current_row.get('selecionado_pcd'),
                        'selecionado_diversidade_racial': current_row.get('selecionado_diversidade_racial'),
                        'selecionado_tipo': current_row.get('selecionado_tipo'),
                        'selecionado_empresa_anterior': current_row.get('selecionado_empresa_anterior'),
                        'selecionado_empregado': current_row.get('selecionado_empregado'),
                        'operacao_status_ultima_proposta': current_row.get('operacao_status_ultima_proposta'),
                        'operacao_selecionado_ultima_proposta': current_row.get('operacao_selecionado_ultima_proposta'),
                        'operacao_motivo_declinio_ultima_proposta': current_row.get('operacao_motivo_declinio_ultima_proposta'),
                    }
                    history_records.append(record)

        prev_data = data
        processed += 1

        if processed % 20 == 0:
            print(f"  Processed {processed} revisions, {len(history_records)} history records...")

    print(f"\nProcessed {processed} revisions, skipped {skipped}")
    print(f"Total history records to insert: {len(history_records)}\n")

    # Insert records
    print("Inserting records into log_history_rebuild...")

    insert_sql = """
        INSERT INTO "RPO_cielo".log_history_rebuild
        (requisicao, status, alterado_por, created_at, updated_at, version,
         cdd_totais_sl, cdd_mulher_sl, cdd_pcd_sl, cdd_diversidade_racial_sl,
         selecionado_nome, selecionado_fonte, selecionado_genero, selecionado_pcd,
         selecionado_diversidade_racial, selecionado_tipo, selecionado_empresa_anterior,
         selecionado_empregado, operacao_status_ultima_proposta,
         operacao_selecionado_ultima_proposta, operacao_motivo_declinio_ultima_proposta)
        VALUES (%s, %s, %s, %s, %s, 1,
                %s, %s, %s, %s,
                %s, %s, %s, %s,
                %s, %s, %s,
                %s, %s,
                %s, %s)
    """

    for record in history_records:
        cur.execute(insert_sql, (
            record['requisicao'],
            record['status'],
            record['alterado_por'],
            record['created_at'],
            record['updated_at'],
            record['cdd_totais_sl'],
            record['cdd_mulher_sl'],
            record['cdd_pcd_sl'],
            record['cdd_diversidade_racial_sl'],
            record['selecionado_nome'],
            record['selecionado_fonte'],
            record['selecionado_genero'],
            record['selecionado_pcd'],
            record['selecionado_diversidade_racial'],
            record['selecionado_tipo'],
            record['selecionado_empresa_anterior'],
            record['selecionado_empregado'],
            record['operacao_status_ultima_proposta'],
            record['operacao_selecionado_ultima_proposta'],
            record['operacao_motivo_declinio_ultima_proposta'],
        ))

    conn.commit()
    print(f"Inserted {len(history_records)} records.\n")

    # Summary
    print("=== Summary ===")

    cur.execute('SELECT COUNT(*) FROM "RPO_cielo".log_history_rebuild')
    total = cur.fetchone()[0]
    print(f"Total rows: {total}")

    cur.execute('SELECT COUNT(DISTINCT requisicao) FROM "RPO_cielo".log_history_rebuild')
    unique_rps = cur.fetchone()[0]
    print(f"Unique RPs: {unique_rps}")

    # Status distribution
    print("\nStatus distribution:")
    cur.execute("""
        SELECT status, COUNT(*) as cnt
        FROM "RPO_cielo".log_history_rebuild
        GROUP BY status
        ORDER BY cnt DESC
    """)
    for status, cnt in cur.fetchall():
        print(f"  {status}: {cnt}")

    # Records with selecionado data
    cur.execute("""
        SELECT COUNT(*) FROM "RPO_cielo".log_history_rebuild
        WHERE selecionado_nome IS NOT NULL AND TRIM(selecionado_nome) != ''
    """)
    with_selecionado = cur.fetchone()[0]
    print(f"\nRecords with selecionado_nome: {with_selecionado}")

    # Records with candidatos data
    cur.execute("""
        SELECT COUNT(*) FROM "RPO_cielo".log_history_rebuild
        WHERE cdd_totais_sl IS NOT NULL AND TRIM(cdd_totais_sl) != ''
    """)
    with_candidatos = cur.fetchone()[0]
    print(f"Records with cdd_totais_sl: {with_candidatos}")

    # Sample record with all fields
    print("\nSample record with selecionado data:")
    cur.execute("""
        SELECT requisicao, status, selecionado_nome, selecionado_fonte,
               cdd_totais_sl, operacao_status_ultima_proposta, created_at
        FROM "RPO_cielo".log_history_rebuild
        WHERE selecionado_nome IS NOT NULL AND TRIM(selecionado_nome) != ''
        LIMIT 1
    """)
    row = cur.fetchone()
    if row:
        print(f"  RP: {row[0]}")
        print(f"  Status: {row[1]}")
        print(f"  Selecionado: {row[2]}")
        print(f"  Fonte: {row[3]}")
        print(f"  Candidatos: {row[4]}")
        print(f"  Status Proposta: {row[5]}")
        print(f"  Date: {row[6]}")

    cur.close()
    conn.close()

    print("\n=== Done! ===")


if __name__ == '__main__':
    main()
