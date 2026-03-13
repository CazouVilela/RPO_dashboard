#!/usr/bin/env python3
"""
Adiciona coluna 'divulgacao' na planilha Appian (.xlsx no Google Drive).

Fluxo:
1. Baixa o .xlsx do Google Drive via service account
2. Lê a coluna 'historico' e parseia datas de divulgação
3. Adiciona coluna 'divulgacao' com datas em vermelho
4. Colore trechos relevantes do histórico em vermelho
5. Re-uploada o .xlsx para o Drive
"""

import re
import os
import sys
from datetime import datetime
from copy import copy

from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload

import openpyxl
from openpyxl.styles import Font, PatternFill

# === CONFIG ===
SERVICE_ACCOUNT_FILE = '/home/cazouvilela/credenciais/gcp_grupohub_service_account.json'
FILE_ID = '1aW71mfAHBSbzIeS9-be9qcpfzTWjzEPf'
SCOPES = ['https://www.googleapis.com/auth/drive']
DOWNLOAD_PATH = '/tmp/appian_vagas.xlsx'
UPLOAD_PATH = '/tmp/appian_vagas_updated.xlsx'

RED_FONT = Font(color='FF0000', bold=True)
RED_FONT_NORMAL = Font(color='FF0000')

# Estratégia: dividir o histórico em entradas (por data DD/MM) e analisar cada uma
# Isso evita que uma entrada capture a data de outra

# Palavras-chave que indicam divulgação numa entrada
DIVULGACAO_KEYWORDS = [
    r'vaga\s+divulgada',
    r'vaga\s+publicada',
    r'divulga[çc][aã]o\s+(?:interna|externa)',
    r'in[ií]cio\s+da\s+divulga[çc][aã]o',
    r'processo\s+seletivo\s+iniciado\s+e\s+vaga\s+divulgada',
    r'descri[çc][aã]o\s+de\s+cargo\s+(?:aprovada|enviada).*?vaga\s+divulgada',
    r'vaga\s+em\s+divulga[çc][aã]o',
    r'liberaram\s+a\s+divulga[çc][aã]o',
]

# Padrão especial: "Vaga publicada na Gupy em DD/MM" (a data está DENTRO do texto, não no início)
GUPY_PATTERN = r'[Vv]aga\s+publicada\s+na\s+Gupy\s+em\s+(\d{1,2}/\d{1,2}(?:/\d{2,4})?)'

# Padrão para "divulgação de DD/MM a DD/MM"
DIVULGACAO_DE_PATTERN = r'divulga[çc][aã]o\s+de\s+(\d{1,2}/\d{1,2}(?:/\d{2,4})?)'


def get_drive_service():
    """Autentica com service account e retorna o serviço do Drive."""
    creds = service_account.Credentials.from_service_account_file(
        SERVICE_ACCOUNT_FILE, scopes=SCOPES
    )
    return build('drive', 'v3', credentials=creds)


def download_xlsx(service):
    """Baixa o .xlsx do Google Drive."""
    print(f"Baixando arquivo {FILE_ID}...")
    request = service.files().get_media(fileId=FILE_ID)
    with open(DOWNLOAD_PATH, 'wb') as f:
        downloader = MediaIoBaseDownload(f, request)
        done = False
        while not done:
            status, done = downloader.next_chunk()
            if status:
                print(f"  Download: {int(status.progress() * 100)}%")
    print(f"  Salvo em {DOWNLOAD_PATH}")
    return DOWNLOAD_PATH


def upload_xlsx(service):
    """Re-uploada o .xlsx atualizado para o Drive."""
    print(f"Uploadando arquivo atualizado...")
    media = MediaFileUpload(UPLOAD_PATH, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    service.files().update(fileId=FILE_ID, media_body=media).execute()
    print(f"  Upload concluído!")


def split_entries(historico):
    """Divide o histórico em entradas individuais por data DD/MM."""
    # Cada entrada começa com DD/MM ou DD/MM/YY(YY)
    # Pode ser separada por \n ou estar colada (sem \n)
    parts = re.split(r'(?:^|\n|(?<=\.))\s*(?=\d{1,2}/\d{1,2}(?:/\d{2,4})?\s*[-–])', historico)
    return [p.strip() for p in parts if p.strip()]


def extract_entry_date(entry):
    """Extrai a data (DD/MM ou DD/MM/YY) do início de uma entrada."""
    m = re.match(r'(\d{1,2}/\d{1,2}(?:/\d{2,4})?)\s*[-–]', entry)
    return m.group(1) if m else None


def extract_divulgacao_date(historico):
    """Extrai a data de divulgação do texto do histórico."""
    if not historico:
        return None, None

    text = str(historico)

    # 1. Prioridade: padrão "Vaga publicada na Gupy em DD/MM" (data explícita no texto)
    gupy_match = re.search(GUPY_PATTERN, text, re.IGNORECASE)
    if gupy_match:
        date_str = gupy_match.group(1)
        # Encontrar a entrada que contém esse trecho
        start = max(0, gupy_match.start() - 100)
        end = min(len(text), gupy_match.end() + 50)
        context = text[start:end]
        return date_str, context.strip()

    # 2. Dividir em entradas e procurar keywords de divulgação
    entries = split_entries(text)
    for entry in entries:
        for kw in DIVULGACAO_KEYWORDS:
            if re.search(kw, entry, re.IGNORECASE):
                entry_date = extract_entry_date(entry)
                if entry_date:
                    return entry_date, entry.strip()

    # 3. Fallback: "divulgação de DD/MM"
    de_match = re.search(DIVULGACAO_DE_PATTERN, text, re.IGNORECASE)
    if de_match:
        return de_match.group(1), text[max(0, de_match.start()-30):de_match.end()+30].strip()

    return None, None


def find_column_by_header(ws, header_name):
    """Encontra a coluna pelo nome do cabeçalho (case-insensitive)."""
    for col in range(1, ws.max_column + 1):
        cell_val = ws.cell(row=1, column=col).value
        if cell_val and str(cell_val).strip().lower() == header_name.lower():
            return col
    return None


def process_workbook(filepath):
    """Processa o workbook: adiciona coluna divulgacao e formata em vermelho."""
    print(f"Abrindo {filepath}...")
    wb = openpyxl.load_workbook(filepath, rich_text=False)

    # Encontrar a aba correta (primeira aba ou a que tem os dados)
    ws = wb.active
    print(f"  Aba ativa: '{ws.title}'")
    print(f"  Dimensões: {ws.dimensions}")
    print(f"  Linhas: {ws.max_row}, Colunas: {ws.max_column}")

    # Listar cabeçalhos
    headers = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        headers.append(val)
        print(f"    Col {col}: {val}")

    # Encontrar coluna 'historico'
    hist_col = find_column_by_header(ws, 'historico')
    if not hist_col:
        # Tentar variações
        for name in ['Historico', 'HISTORICO', 'Histórico', 'histórico']:
            hist_col = find_column_by_header(ws, name)
            if hist_col:
                break

    if not hist_col:
        print("ERRO: Coluna 'historico' não encontrada!")
        print(f"  Cabeçalhos encontrados: {headers}")
        sys.exit(1)

    print(f"  Coluna 'historico' encontrada na posição {hist_col}")

    # Encontrar coluna 'id'
    id_col = find_column_by_header(ws, 'id')
    if not id_col:
        for name in ['ID', 'Id']:
            id_col = find_column_by_header(ws, name)
            if id_col:
                break

    # Verificar se coluna 'divulgacao' já existe
    divulgacao_col = find_column_by_header(ws, 'divulgacao')
    if not divulgacao_col:
        for name in ['Divulgacao', 'DIVULGACAO', 'Divulgação', 'divulgação']:
            divulgacao_col = find_column_by_header(ws, name)
            if divulgacao_col:
                break

    if divulgacao_col:
        print(f"  Coluna 'divulgacao' já existe na posição {divulgacao_col}. Será atualizada.")
    else:
        # Adicionar coluna após historico
        divulgacao_col = ws.max_column + 1
        ws.cell(row=1, column=divulgacao_col, value='divulgacao')
        ws.cell(row=1, column=divulgacao_col).font = Font(bold=True)
        print(f"  Coluna 'divulgacao' criada na posição {divulgacao_col}")

    # Processar cada linha
    found_count = 0
    total_count = 0

    for row in range(2, ws.max_row + 1):
        hist_cell = ws.cell(row=row, column=hist_col)
        historico = hist_cell.value
        row_id = ws.cell(row=row, column=id_col).value if id_col else row - 1

        if not historico:
            continue

        total_count += 1
        date_str, line = extract_divulgacao_date(str(historico))

        if date_str:
            found_count += 1
            # Preencher data em vermelho
            div_cell = ws.cell(row=row, column=divulgacao_col, value=date_str)
            div_cell.font = RED_FONT

            # Colorir célula do histórico com fonte vermelha onde tem divulgação
            # Como openpyxl não suporta rich text parcial facilmente,
            # vamos colorir toda a célula de histórico com um fundo sutil
            # para indicar que tem divulgação
            hist_cell.font = RED_FONT_NORMAL

            print(f"  ID {row_id}: {date_str} <- \"{line[:80]}...\"" if line and len(line) > 80 else f"  ID {row_id}: {date_str} <- \"{line}\"")
        else:
            # Limpar célula de divulgacao se não encontrou
            ws.cell(row=row, column=divulgacao_col, value=None)

    print(f"\n  Resultado: {found_count}/{total_count} registros com divulgação identificada")

    # Salvar
    wb.save(UPLOAD_PATH)
    print(f"  Salvo em {UPLOAD_PATH}")
    return wb, found_count


def main():
    print("=" * 60)
    print("APPIAN - Processamento coluna Divulgação")
    print("=" * 60)

    # 1. Conectar ao Drive
    service = get_drive_service()

    # 2. Baixar xlsx
    download_xlsx(service)

    # 3. Processar
    wb, count = process_workbook(DOWNLOAD_PATH)

    if count == 0:
        print("\nNenhuma divulgação encontrada. Abortando upload.")
        sys.exit(1)

    # 4. Confirmar upload
    print(f"\n{'=' * 60}")
    print(f"Pronto para upload: {count} datas de divulgação encontradas")

    if '--dry-run' in sys.argv:
        print("  [DRY RUN] Upload não realizado.")
        print(f"  Arquivo disponível em: {UPLOAD_PATH}")
        return

    if '--yes' not in sys.argv:
        resp = input("Prosseguir com upload? (s/N): ")
        if resp.lower() not in ('s', 'sim', 'y', 'yes'):
            print("Upload cancelado.")
            print(f"  Arquivo disponível em: {UPLOAD_PATH}")
            return

    # 5. Upload
    upload_xlsx(service)
    print(f"\n{'=' * 60}")
    print("Concluído com sucesso!")


if __name__ == '__main__':
    main()
